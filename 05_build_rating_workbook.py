"""
NFL Injury Risk Project - Excel Rating Factor Workbook
=========================================================
Builds the Phase 3 deliverable: an Excel workbook containing raw
player-season data, a formula-driven pivot-style summary, a rating
factor table (position + age band relativities), and a LOOKUP-driven
rate lookup tool -- the "rate manual" a real actuarial pricing
exercise would produce.

Run: python 05_build_rating_workbook.py
Reads:  ../data/player_season_flat.csv
Writes: ../excel/nfl_injury_rating_workbook.xlsx
"""

import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

CSV_PATH = "../data/player_season_flat.csv"
OUT_PATH = "../excel/nfl_injury_rating_workbook.xlsx"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
INPUT_FONT = Font(name=FONT_NAME, color="0000FF")       # blue = hardcoded input, per convention
FORMULA_FONT = Font(name=FONT_NAME, color="000000")     # black = formula
ASSUMPTION_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=9, color="595959")
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

POSITIONS = ["RB", "OL", "QB"]
AGE_BANDS = ["<=24", "25-27", "28-30", "31+"]


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_raw_data_sheet(wb):
    ws = wb.create_sheet("Player_Season_Data")
    headers = [
        "Position", "Season", "Player_ID", "Player_Name", "Age_Band",
        "Claim_Classification", "Games_Missed", "Is_PTD", "Is_TTD",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    with open(CSV_PATH) as f:
        reader = csv.DictReader(f)
        row_i = 2
        for row in reader:
            games_missed = row["games_missed"] if row["games_missed"] else ""
            ws.append([
                row["position_group"],
                int(row["season"]),
                row["player_id"],
                row["player_name"],
                row["age_band"],
                row["claim_classification"],
                float(games_missed) if games_missed != "" else None,
                None,  # placeholder, formula written below
                None,
            ])
            ws.cell(row=row_i, column=8).value = f'=IF(F{row_i}="PTD",1,0)'
            ws.cell(row=row_i, column=9).value = f'=IF(F{row_i}="TTD",1,0)'
            for c in [1, 2, 3, 4, 5, 6, 7]:
                ws.cell(row=row_i, column=c).font = INPUT_FONT
            for c in [8, 9]:
                ws.cell(row=row_i, column=c).font = FORMULA_FONT
            row_i += 1

    autofit(ws, [10, 9, 13, 20, 10, 20, 13, 8, 8])
    ws.freeze_panes = "A2"
    last_row = row_i - 1
    return ws, last_row


def build_summary_pivot_sheet(wb, data_last_row):
    ws = wb.create_sheet("Summary_Pivot")
    ws["A1"] = "Formula-driven pivot-style summary (COUNTIFS/AVERAGEIFS against Player_Season_Data)"
    ws["A1"].font = NOTE_FONT
    ws["A2"] = (
        "Note: openpyxl cannot reliably create native Excel PivotTable objects, so this "
        "sheet reproduces the same output using SUMIFS/COUNTIFS formulas instead. To build "
        "a native PivotTable from the same source, select Player_Season_Data and use "
        "Insert > PivotTable in Excel."
    )
    ws["A2"].font = NOTE_FONT
    ws.merge_cells("A2:I2")
    ws["A3"] = (
        "Exposure definition note: 'Exposure' here excludes rows tagged 'no_data' "
        "(rostered players with no matching performance/injury record -- see "
        "LIMITATIONS.md). This differs slightly from the SQL layer's 'broad_exposure' "
        "(sql/01_frequency_by_position.sql), which does not exclude those rows -- so "
        "totals here will not match that query exactly. Both are valid, disclosed choices."
    )
    ws["A3"].font = NOTE_FONT
    ws.merge_cells("A3:I3")

    headers = [
        "Position", "Age_Band", "Exposure", "PTD_Claims", "TTD_Claims",
        "PTD_Frequency", "TTD_Frequency", "PTD_Avg_Games_Missed", "TTD_Avg_Games_Missed",
    ]
    header_row = 5
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i, value=h)
    style_header_row(ws, header_row, len(headers))

    data_range = f"Player_Season_Data!$A$2:$A${data_last_row}"
    age_range = f"Player_Season_Data!$E$2:$E${data_last_row}"
    class_range = f"Player_Season_Data!$F$2:$F${data_last_row}"
    games_range = f"Player_Season_Data!$G$2:$G${data_last_row}"

    r = header_row + 1
    for pos in POSITIONS:
        for band in AGE_BANDS:
            ws.cell(row=r, column=1, value=pos).font = INPUT_FONT
            ws.cell(row=r, column=2, value=band).font = INPUT_FONT

            # Exposure = all rows for this position+age_band except pure "no_data" rows
            # NOTE: age-band criteria are escaped as ("="&$B{r}) because Excel/LibreOffice
            # COUNTIFS/SUMIFS interpret a criteria value beginning with <, >, <=, >=, <>
            # as a comparison operator, not literal text -- "<=24" would otherwise be read
            # as "<= 24" (a numeric comparison) instead of matching the text label.
            ws.cell(row=r, column=3).value = (
                f'=COUNTIFS({data_range},$A{r},{age_range},"="&$B{r})'
                f'-COUNTIFS({data_range},$A{r},{age_range},"="&$B{r},{class_range},"no_data")'
            )
            ws.cell(row=r, column=4).value = (
                f'=COUNTIFS({data_range},$A{r},{age_range},"="&$B{r},{class_range},"PTD")'
            )
            ws.cell(row=r, column=5).value = (
                f'=COUNTIFS({data_range},$A{r},{age_range},"="&$B{r},{class_range},"TTD")'
            )
            ws.cell(row=r, column=6).value = f'=IFERROR(D{r}/C{r},0)'
            ws.cell(row=r, column=7).value = f'=IFERROR(E{r}/C{r},0)'
            ws.cell(row=r, column=8).value = (
                f'=IFERROR(AVERAGEIFS({games_range},{data_range},$A{r},{age_range},"="&$B{r},'
                f'{class_range},"PTD"),0)'
            )
            ws.cell(row=r, column=9).value = (
                f'=IFERROR(AVERAGEIFS({games_range},{data_range},$A{r},{age_range},"="&$B{r},'
                f'{class_range},"TTD"),0)'
            )
            for c in [3, 4, 5, 6, 7, 8, 9]:
                ws.cell(row=r, column=c).font = FORMULA_FONT
            r += 1

    for row in ws.iter_rows(min_row=header_row, max_row=r - 1, min_col=1, max_col=9):
        for cell in row:
            cell.border = THIN_BORDER

    # Helper key column for simple (non-array) INDEX/MATCH lookups elsewhere in the workbook
    ws.cell(row=header_row, column=10, value="Lookup_Key")
    ws.cell(row=header_row, column=10).fill = HEADER_FILL
    ws.cell(row=header_row, column=10).font = HEADER_FONT
    for row_n in range(header_row + 1, r):
        cell = ws.cell(row=row_n, column=10)
        cell.value = f'=A{row_n}&"|"&B{row_n}'
        cell.font = FORMULA_FONT
        cell.border = THIN_BORDER

    # number formats
    for row in range(header_row + 1, r):
        ws.cell(row=row, column=6).number_format = "0.0%"
        ws.cell(row=row, column=7).number_format = "0.0%"
        ws.cell(row=row, column=8).number_format = "0.00"
        ws.cell(row=row, column=9).number_format = "0.00"

    autofit(ws, [10, 10, 10, 11, 11, 13, 13, 18, 18, 14])
    return ws, header_row + 1, r - 1  # first/last data row


def build_rating_factors_sheet(wb, summary_first, summary_last):
    ws = wb.create_sheet("Rating_Factors")
    ws["A1"] = "Rating Factors -- base rate + relativities (relative to OL / 25-27 as the base class)"
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=12)
    ws.merge_cells("A1:F1")

    ws["A3"] = "Base class:"
    ws["B3"] = "OL"
    ws["C3"] = "25-27"
    ws["A3"].font = FORMULA_FONT
    ws["B3"].fill = ASSUMPTION_FILL
    ws["C3"].fill = ASSUMPTION_FILL
    ws["B3"].font = INPUT_FONT
    ws["C3"].font = INPUT_FONT
    ws["D3"] = (
        "<- base position and age band chosen as the highest-volume, most stable cell "
        "in the portfolio, not an arbitrary pick"
    )
    ws["D3"].font = NOTE_FONT
    ws.merge_cells("D3:H3")

    ws["A5"] = "Base PTD frequency:"
    ws["A6"] = "Base TTD frequency:"
    # NOTE: age-band criteria ($C$3, and $A{r} below) are escaped as ("="&cell) because
    # Excel/LibreOffice SUMIFS/COUNTIFS interpret a criteria value beginning with
    # <, >, <=, >=, <> as a comparison operator, not literal text -- "<=24" would
    # otherwise be read as "<= 24" (numeric) instead of matching the text label.
    ws["B5"] = (
        f'=SUMIFS(Summary_Pivot!$D${summary_first}:$D${summary_last},'
        f'Summary_Pivot!$A${summary_first}:$A${summary_last},$B$3,'
        f'Summary_Pivot!$B${summary_first}:$B${summary_last},"="&$C$3)'
        f'/SUMIFS(Summary_Pivot!$C${summary_first}:$C${summary_last},'
        f'Summary_Pivot!$A${summary_first}:$A${summary_last},$B$3,'
        f'Summary_Pivot!$B${summary_first}:$B${summary_last},"="&$C$3)'
    )
    ws["B6"] = (
        f'=SUMIFS(Summary_Pivot!$E${summary_first}:$E${summary_last},'
        f'Summary_Pivot!$A${summary_first}:$A${summary_last},$B$3,'
        f'Summary_Pivot!$B${summary_first}:$B${summary_last},"="&$C$3)'
        f'/SUMIFS(Summary_Pivot!$C${summary_first}:$C${summary_last},'
        f'Summary_Pivot!$A${summary_first}:$A${summary_last},$B$3,'
        f'Summary_Pivot!$B${summary_first}:$B${summary_last},"="&$C$3)'
    )
    ws["B5"].number_format = "0.00%"
    ws["B6"].number_format = "0.00%"
    for cell in ["B5", "B6"]:
        ws[cell].font = FORMULA_FONT

    # ---- Position relativities (holding age band = base) ----
    ws["A9"] = "Position Relativities (at base age band)"
    ws["A9"].font = Font(name=FONT_NAME, bold=True)
    pos_headers = ["Position", "PTD_Frequency", "TTD_Frequency", "PTD_Relativity", "TTD_Relativity"]
    for i, h in enumerate(pos_headers, start=1):
        ws.cell(row=10, column=i, value=h)
    style_header_row(ws, 10, len(pos_headers))

    r = 11
    for pos in POSITIONS:
        ws.cell(row=r, column=1, value=pos).font = INPUT_FONT
        ws.cell(row=r, column=2).value = (
            f'=SUMIFS(Summary_Pivot!$D${summary_first}:$D${summary_last},'
            f'Summary_Pivot!$A${summary_first}:$A${summary_last},$A{r},'
            f'Summary_Pivot!$B${summary_first}:$B${summary_last},"="&$C$3)'
            f'/SUMIFS(Summary_Pivot!$C${summary_first}:$C${summary_last},'
            f'Summary_Pivot!$A${summary_first}:$A${summary_last},$A{r},'
            f'Summary_Pivot!$B${summary_first}:$B${summary_last},"="&$C$3)'
        )
        ws.cell(row=r, column=3).value = (
            f'=SUMIFS(Summary_Pivot!$E${summary_first}:$E${summary_last},'
            f'Summary_Pivot!$A${summary_first}:$A${summary_last},$A{r},'
            f'Summary_Pivot!$B${summary_first}:$B${summary_last},"="&$C$3)'
            f'/SUMIFS(Summary_Pivot!$C${summary_first}:$C${summary_last},'
            f'Summary_Pivot!$A${summary_first}:$A${summary_last},$A{r},'
            f'Summary_Pivot!$B${summary_first}:$B${summary_last},"="&$C$3)'
        )
        ws.cell(row=r, column=4).value = f'=B{r}/$B$5'
        ws.cell(row=r, column=5).value = f'=C{r}/$B$6'
        for c in [2, 3]:
            ws.cell(row=r, column=c).number_format = "0.00%"
        for c in [4, 5]:
            ws.cell(row=r, column=c).number_format = "0.00"
            ws.cell(row=r, column=c).font = FORMULA_FONT
        ws.cell(row=r, column=2).font = FORMULA_FONT
        ws.cell(row=r, column=3).font = FORMULA_FONT
        r += 1
    pos_table_last = r - 1

    # ---- Age band relativities (holding position = base) ----
    r += 2
    ws.cell(row=r, column=1, value="Age Band Relativities (at base position)").font = Font(
        name=FONT_NAME, bold=True
    )
    r += 1
    age_header_row = r
    for i, h in enumerate(["Age_Band", "PTD_Frequency", "TTD_Frequency", "PTD_Relativity", "TTD_Relativity"], start=1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, 5)
    r += 1
    age_table_first = r
    for band in AGE_BANDS:
        ws.cell(row=r, column=1, value=band).font = INPUT_FONT
        ws.cell(row=r, column=2).value = (
            f'=SUMIFS(Summary_Pivot!$D${summary_first}:$D${summary_last},'
            f'Summary_Pivot!$B${summary_first}:$B${summary_last},"="&$A{r},'
            f'Summary_Pivot!$A${summary_first}:$A${summary_last},$B$3)'
            f'/SUMIFS(Summary_Pivot!$C${summary_first}:$C${summary_last},'
            f'Summary_Pivot!$B${summary_first}:$B${summary_last},"="&$A{r},'
            f'Summary_Pivot!$A${summary_first}:$A${summary_last},$B$3)'
        )
        ws.cell(row=r, column=3).value = (
            f'=SUMIFS(Summary_Pivot!$E${summary_first}:$E${summary_last},'
            f'Summary_Pivot!$B${summary_first}:$B${summary_last},"="&$A{r},'
            f'Summary_Pivot!$A${summary_first}:$A${summary_last},$B$3)'
            f'/SUMIFS(Summary_Pivot!$C${summary_first}:$C${summary_last},'
            f'Summary_Pivot!$B${summary_first}:$B${summary_last},"="&$A{r},'
            f'Summary_Pivot!$A${summary_first}:$A${summary_last},$B$3)'
        )
        ws.cell(row=r, column=4).value = f'=B{r}/$B$5'
        ws.cell(row=r, column=5).value = f'=C{r}/$B$6'
        for c in [2, 3]:
            ws.cell(row=r, column=c).number_format = "0.00%"
            ws.cell(row=r, column=c).font = FORMULA_FONT
        for c in [4, 5]:
            ws.cell(row=r, column=c).number_format = "0.00"
            ws.cell(row=r, column=c).font = FORMULA_FONT
        r += 1
    age_table_last = r - 1

    autofit(ws, [16, 15, 15, 15, 15, 40])
    return ws, pos_table_last, age_table_first, age_table_last


def build_rate_lookup_sheet(wb, summary_first, summary_last):
    ws = wb.create_sheet("Rate_Lookup_Tool")
    ws["A1"] = "Rate Lookup Tool -- select a Position and Age Band to see the indicated rate"
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=12)
    ws.merge_cells("A1:E1")

    ws["A3"] = "Position:"
    ws["A4"] = "Age Band:"
    ws["B3"] = "RB"
    ws["B4"] = "25-27"
    for cell in ["B3", "B4"]:
        ws[cell].fill = ASSUMPTION_FILL
        ws[cell].font = INPUT_FONT
    ws["C3"] = "<- edit these two cells (dropdowns) to look up a rate"
    ws["C3"].font = NOTE_FONT

    dv_pos = DataValidation(type="list", formula1=f'"{",".join(POSITIONS)}"', allow_blank=False)
    dv_age = DataValidation(type="list", formula1=f'"{",".join(AGE_BANDS)}"', allow_blank=False)
    ws.add_data_validation(dv_pos)
    ws.add_data_validation(dv_age)
    dv_pos.add(ws["B3"])
    dv_age.add(ws["B4"])

    labels = [
        ("Exposure (player-seasons)", "C"),
        ("PTD Frequency", "F"),
        ("TTD Frequency", "G"),
        ("PTD Avg Games Missed", "H"),
        ("TTD Avg Games Missed", "I"),
    ]
    key_range = f"Summary_Pivot!$J${summary_first}:$J${summary_last}"

    r = 6
    for label, col_letter in labels:
        ws.cell(row=r, column=1, value=label).font = FORMULA_FONT
        cell = ws.cell(row=r, column=2)
        cell.value = (
            f'=INDEX(Summary_Pivot!${col_letter}${summary_first}:${col_letter}${summary_last},'
            f"MATCH($B$3&\"|\"&$B$4,{key_range},0))"
        )
        cell.font = FORMULA_FONT
        if col_letter in ("F", "G"):
            cell.number_format = "0.00%"
        elif col_letter in ("H", "I"):
            cell.number_format = "0.00"
        r += 1

    ws["A12"] = "Expected games lost per player-season (PTD+TTD combined):"
    ws["A12"].font = Font(name=FONT_NAME, bold=True)
    ws["B12"] = "=(B7*B9)+(B8*B10)"
    ws["B12"].font = Font(name=FONT_NAME, bold=True)
    ws["B12"].number_format = "0.00"

    ws["A14"] = (
        "Note: this is a frequency x severity pure-premium-style metric measured in games "
        "lost, not dollars. Dollar-based pricing (guaranteed money at risk) is built out in "
        "the Phase 5 pricing memo, using sql/04_dollar_exposure_by_classification.sql as the "
        "starting point."
    )
    ws["A14"].font = NOTE_FONT
    ws.merge_cells("A14:F14")

    autofit(ws, [42, 14, 45])
    return ws


def main():
    Path("../excel").mkdir(exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)  # drop default empty sheet

    raw_ws, data_last_row = build_raw_data_sheet(wb)
    print(f"Player_Season_Data: {data_last_row - 1:,} rows written")

    summary_ws, summary_first, summary_last = build_summary_pivot_sheet(wb, data_last_row)
    print(f"Summary_Pivot: rows {summary_first}-{summary_last}")

    build_rating_factors_sheet(wb, summary_first, summary_last)
    print("Rating_Factors sheet built")

    build_rate_lookup_sheet(wb, summary_first, summary_last)
    print("Rate_Lookup_Tool sheet built")

    wb.save(OUT_PATH)
    print(f"Saved workbook to {OUT_PATH}")


if __name__ == "__main__":
    main()
